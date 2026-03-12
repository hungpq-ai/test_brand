#!/usr/bin/env python3
"""
Calculate Aggregate AI Visibility Scores from results CSV
Implements the full scoring framework across all prompts per brand per engine

Usage: python3 calculate_aggregate_scores.py results.csv
"""
import sys
import pandas as pd
from pathlib import Path
from scoring import BrandScore, calculate_ai_visibility_score, get_score_grade


def calculate_aggregates(csv_path):
    """Calculate aggregate AI Visibility Scores"""

    print(f"\n{'='*100}")
    print(f"Calculating Aggregate AI Visibility Scores")
    print(f"{'='*100}\n")
    print(f"Input: {csv_path}\n")

    # Load results
    df = pd.read_csv(csv_path)

    # Required columns
    required_cols = ['AI Engine', 'Brand', 'Mention', 'Rank', 'Citation Type']
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        print(f"❌ Missing columns: {missing}")
        return

    print(f"Total rows: {len(df)}")
    print(f"Unique brands: {df['Brand'].nunique()}")
    print(f"Unique engines: {df['AI Engine'].nunique()}")
    print(f"Unique queries: {df['Query'].nunique()}\n")

    # Aggregate by Engine + Brand
    results = []

    for engine in df['AI Engine'].unique():
        for brand in df['Brand'].unique():
            # Filter data for this engine + brand
            subset = df[(df['AI Engine'] == engine) & (df['Brand'] == brand)]

            if len(subset) == 0:
                continue

            # Create BrandScore objects
            brand_scores = []
            for _, row in subset.iterrows():
                brand_scores.append(BrandScore(
                    brand=brand,
                    mentioned=(row['Mention'] == 'Yes'),
                    rank=int(row['Rank']) if pd.notna(row['Rank']) else None,
                    citation_type=row['Citation Type'] if pd.notna(row['Citation Type']) else 'none'
                ))

            # Calculate aggregate scores
            scores = calculate_ai_visibility_score(brand_scores)

            # Get grade
            grade = get_score_grade(scores['ai_visibility_score'])

            results.append({
                'AI Engine': engine,
                'Brand': brand,
                'Total Prompts': scores['total_prompts'],
                'Mention Count': scores['mention_count'],
                'Mention Rate %': scores['mention_rate'],
                'Brand Mention Score': scores['mention_score'],
                'Ranking Score': scores['ranking_score'],
                'Citation Score': scores['citation_score'],
                'AI Visibility Score': scores['ai_visibility_score'],
                'Grade': grade,
            })

    # Create results DataFrame
    results_df = pd.DataFrame(results)

    # Sort by AI Visibility Score descending
    results_df = results_df.sort_values('AI Visibility Score', ascending=False)

    # Save to CSV
    output_path = csv_path.replace('.csv', '_AGGREGATED.csv')
    results_df.to_csv(output_path, index=False)

    print(f"{'='*100}")
    print(f"Aggregate Scores Calculated!")
    print(f"{'='*100}\n")
    print(f"Output saved to: {output_path}\n")

    # Display top results
    print("Top 10 Results by AI Visibility Score:")
    print(results_df.head(10).to_string(index=False))

    print(f"\n{'='*100}")
    print("Summary by Brand (across all engines):")
    print(f"{'='*100}\n")

    brand_summary = results_df.groupby('Brand').agg({
        'AI Visibility Score': 'mean',
        'Mention Rate %': 'mean',
        'Ranking Score': 'mean',
        'Citation Score': 'mean',
    }).round(2).sort_values('AI Visibility Score', ascending=False)

    print(brand_summary)

    print(f"\n{'='*100}")
    print("Summary by Engine (across all brands):")
    print(f"{'='*100}\n")

    engine_summary = results_df.groupby('AI Engine').agg({
        'AI Visibility Score': 'mean',
        'Mention Rate %': 'mean',
        'Ranking Score': 'mean',
        'Citation Score': 'mean',
    }).round(2).sort_values('AI Visibility Score', ascending=False)

    print(engine_summary)

    return results_df


def create_excel_with_aggregates(csv_path):
    """Create Excel file with both detailed and aggregate results"""

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Calculate aggregates
    agg_df = calculate_aggregates(csv_path)

    # Load original results
    detail_df = pd.read_csv(csv_path)

    # Create Excel
    excel_path = csv_path.replace('.csv', '_WITH_AGGREGATES.xlsx')

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Sheet 1: Aggregate Scores
        agg_df.to_excel(writer, sheet_name='AI Visibility Scores', index=False)

        # Sheet 2: Detailed Results
        detail_df.to_excel(writer, sheet_name='Detailed Results', index=False)

    # Format the Excel
    wb = load_workbook(excel_path)

    # Format Aggregate sheet
    ws_agg = wb['AI Visibility Scores']

    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for cell in ws_agg[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Color-code grades
    grade_colors = {
        'S+': 'C6EFCE',  # Light green
        'S': 'C6EFCE',
        'A': 'E2EFDA',   # Very light green
        'B': 'FFF2CC',   # Light yellow
        'C': 'FFE699',   # Yellow
        'D': 'FFEB9C',   # Orange-yellow
        'F': 'FFC7CE',   # Light red
    }

    grade_col = None
    for col_idx, cell in enumerate(ws_agg[1], 1):
        if cell.value == 'Grade':
            grade_col = col_idx
            break

    if grade_col:
        for row in ws_agg.iter_rows(min_row=2, max_row=ws_agg.max_row, min_col=grade_col, max_col=grade_col):
            cell = row[0]
            if cell.value in grade_colors:
                cell.fill = PatternFill(start_color=grade_colors[cell.value], end_color=grade_colors[cell.value], fill_type="solid")
                cell.font = Font(bold=True, size=12)

    # Column widths
    ws_agg.column_dimensions['A'].width = 12  # Engine
    ws_agg.column_dimensions['B'].width = 12  # Brand
    ws_agg.column_dimensions['C'].width = 15  # Total Prompts
    ws_agg.column_dimensions['D'].width = 15  # Mention Count
    ws_agg.column_dimensions['E'].width = 15  # Mention Rate
    ws_agg.column_dimensions['F'].width = 20  # Brand Mention Score
    ws_agg.column_dimensions['G'].width = 15  # Ranking Score
    ws_agg.column_dimensions['H'].width = 15  # Citation Score
    ws_agg.column_dimensions['I'].width = 20  # AI Visibility Score
    ws_agg.column_dimensions['J'].width = 10  # Grade

    # Freeze panes
    ws_agg.freeze_panes = 'C2'

    wb.save(excel_path)

    print(f"\n✅ Excel file created: {excel_path}")
    print(f"   - Sheet 1: AI Visibility Scores (Aggregated)")
    print(f"   - Sheet 2: Detailed Results (Per-prompt)\n")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 calculate_aggregate_scores.py <results.csv>")
        print("\nExample:")
        print("  python3 calculate_aggregate_scores.py output/results_20260312_133612.csv")
        print("\nThis will create:")
        print("  1. results_20260312_133612_AGGREGATED.csv")
        print("  2. results_20260312_133612_WITH_AGGREGATES.xlsx (with formatting)")
        sys.exit(1)

    csv_path = sys.argv[1]

    if not Path(csv_path).exists():
        print(f"❌ File not found: {csv_path}")
        sys.exit(1)

    # Calculate and save aggregates
    create_excel_with_aggregates(csv_path)
